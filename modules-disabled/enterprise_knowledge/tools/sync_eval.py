#!/usr/bin/env python
"""Regenerate access/public_evaluation.csv from the workspace dataset xlsx.

Usage: python tools/sync_eval.py "<path to ai_workspace_dataset_...xlsx>"
Reads the Public_Evaluation sheet and writes the columns the eval harness needs.
"""
from __future__ import annotations

import csv
import re
import sys
from pathlib import Path

_QUESTION_ID_RE = re.compile(r"^P\d+$")

_COLS = [
    "question_id",
    "user_id",
    "user_role",
    "user_department",
    "question_vi",
    "expected_permission",
    "expected_document_id",
]


def main(argv: list[str]) -> int:
    if len(argv) != 2:
        print("usage: sync_eval.py <dataset.xlsx>", file=sys.stderr)
        return 2
    from openpyxl import load_workbook

    ws = load_workbook(argv[1], read_only=True, data_only=True)["Public_Evaluation"]
    all_rows = list(ws.iter_rows(values_only=True))

    # Locate the header row by its exact first cell — a leading sheet-title row
    # (e.g. "Public Evaluation") also starts with "P" and must not be mistaken
    # for either the header or a data row.
    header_row = next(r for r in all_rows if r and r[0] == "question_id")
    header = [str(c) for c in header_row]
    idx = {name: header.index(name) for name in _COLS if name in header}

    # Data rows are identified by a strict "P<digits>" question_id, not merely
    # a value that starts with "P".
    data_rows = [r for r in all_rows if r and r[0] is not None and _QUESTION_ID_RE.match(str(r[0]))]

    out_path = Path(__file__).resolve().parent.parent / "access" / "public_evaluation.csv"
    with open(out_path, "w", newline="", encoding="utf-8") as fh:
        writer = csv.writer(fh)
        writer.writerow(_COLS)
        for r in data_rows:
            writer.writerow([r[idx[name]] if name in idx else "" for name in _COLS])
    print(f"wrote {out_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main(sys.argv))
