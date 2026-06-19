#!/usr/bin/env python
"""CSV/Excel loader, profiler, and insight extractor for the data_analysis module.

Subcommands: profile, head, describe, insights, list-files.

Reads tabular data from .csv / .tsv / .xlsx / .xls and emits compact JSON
the agent can reason over. pandas + openpyxl are optional — falls back to
csv stdlib for plain CSV so the module is usable without extras.
"""

from __future__ import annotations

import argparse
import csv
import json
import math
import sys
from pathlib import Path
from typing import Any

ROOT = Path(__file__).resolve().parent.parent
DATA_DIR = ROOT / "data"


# ─── Loading ─────────────────────────────────────────────────────────────

def _resolve(path: str) -> Path:
    p = Path(path).expanduser()
    if not p.is_absolute():
        # Try relative to module data dir, then CWD.
        cand = DATA_DIR / p
        if cand.exists():
            return cand
        p = Path.cwd() / p
    return p


def _load_csv_stdlib(path: Path, delim: str | None = None) -> tuple[list[str], list[list[str]]]:
    with path.open("r", newline="", encoding="utf-8-sig") as fh:
        sample = fh.read(4096)
        fh.seek(0)
        if delim is None:
            try:
                dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
                delim = dialect.delimiter
            except csv.Error:
                delim = ","
        reader = csv.reader(fh, delimiter=delim)
        rows = list(reader)
    if not rows:
        return [], []
    return rows[0], rows[1:]


def _load_excel(path: Path, sheet: str | int | None) -> tuple[list[str], list[list[str]]]:
    try:
        from openpyxl import load_workbook  # type: ignore
    except ImportError as exc:
        raise SystemExit(
            "ERROR: openpyxl not installed. `pip install openpyxl` to read .xlsx"
        ) from exc
    wb = load_workbook(filename=str(path), read_only=True, data_only=True)
    if sheet is None:
        ws = wb.active
    elif isinstance(sheet, int):
        ws = wb.worksheets[sheet]
    else:
        ws = wb[sheet]
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header = [str(c) if c is not None else "" for c in next(rows_iter)]
    except StopIteration:
        return [], []
    rows = [["" if c is None else str(c) for c in r] for r in rows_iter]
    return header, rows


def load(path: str, sheet: str | int | None = None) -> tuple[list[str], list[list[str]], Path]:
    p = _resolve(path)
    if not p.exists():
        raise SystemExit(f"ERROR: file not found: {p}")
    suffix = p.suffix.lower()
    if suffix in (".xlsx", ".xlsm", ".xls"):
        header, rows = _load_excel(p, sheet)
    elif suffix in (".tsv",):
        header, rows = _load_csv_stdlib(p, delim="\t")
    else:
        header, rows = _load_csv_stdlib(p)
    return header, rows, p


# ─── Profiling ───────────────────────────────────────────────────────────

def _is_number(s: str) -> bool:
    if s is None or s == "":
        return False
    try:
        float(s.replace(",", "")) if "," in s and s.count(",") == 1 else float(s)
        return True
    except (ValueError, AttributeError):
        return False


def _to_num(s: str) -> float | None:
    if s is None or s == "":
        return None
    try:
        return float(s)
    except ValueError:
        try:
            return float(s.replace(",", ""))
        except ValueError:
            return None


def _guess_type(values: list[str]) -> str:
    nonempty = [v for v in values if v not in (None, "")]
    if not nonempty:
        return "empty"
    num_count = sum(1 for v in nonempty if _is_number(v))
    if num_count / len(nonempty) >= 0.9:
        return "number"
    # date heuristic
    date_hits = sum(1 for v in nonempty[:200] if _looks_like_date(v))
    if date_hits / min(len(nonempty), 200) >= 0.7:
        return "date"
    return "string"


def _looks_like_date(s: str) -> bool:
    s = s.strip()
    if len(s) < 6 or len(s) > 32:
        return False
    has_sep = ("-" in s or "/" in s) or (":" in s)
    has_digit = any(c.isdigit() for c in s)
    return has_sep and has_digit


def _num_stats(nums: list[float]) -> dict[str, float]:
    if not nums:
        return {"count": 0}
    nums_sorted = sorted(nums)
    n = len(nums)
    mean = sum(nums) / n
    var = sum((x - mean) ** 2 for x in nums) / n
    std = math.sqrt(var)

    def _pct(p: float) -> float:
        if n == 1:
            return nums_sorted[0]
        k = (n - 1) * p
        f = math.floor(k)
        c = math.ceil(k)
        return nums_sorted[f] if f == c else nums_sorted[f] + (nums_sorted[c] - nums_sorted[f]) * (k - f)

    return {
        "count": n,
        "min": nums_sorted[0],
        "p25": _pct(0.25),
        "median": _pct(0.5),
        "p75": _pct(0.75),
        "max": nums_sorted[-1],
        "mean": mean,
        "std": std,
        "sum": sum(nums),
    }


def profile_columns(header: list[str], rows: list[list[str]]) -> list[dict[str, Any]]:
    cols: list[dict[str, Any]] = []
    for i, name in enumerate(header):
        column = [r[i] if i < len(r) else "" for r in rows]
        nonempty = [v for v in column if v not in (None, "")]
        kind = _guess_type(column)
        info: dict[str, Any] = {
            "name": name,
            "type": kind,
            "n": len(column),
            "non_null": len(nonempty),
            "null": len(column) - len(nonempty),
            "unique": len(set(nonempty)),
        }
        if kind == "number":
            nums = [n for n in (_to_num(v) for v in nonempty) if n is not None]
            info["stats"] = _num_stats(nums)
        else:
            from collections import Counter
            top = Counter(nonempty).most_common(5)
            info["top_values"] = [{"value": v, "count": c} for v, c in top]
        cols.append(info)
    return cols


# ─── Insights ────────────────────────────────────────────────────────────

def derive_insights(header: list[str], rows: list[list[str]], profile: list[dict[str, Any]]) -> list[dict[str, Any]]:
    insights: list[dict[str, Any]] = []
    n = len(rows)
    insights.append({"kind": "shape", "text": f"Dataset has {n:,} rows × {len(header)} columns."})

    # Missing data hot-spots
    for col in profile:
        if col["n"] == 0:
            continue
        ratio = col["null"] / col["n"]
        if ratio >= 0.3:
            insights.append({
                "kind": "missing",
                "severity": "high" if ratio >= 0.6 else "medium",
                "column": col["name"],
                "text": f"`{col['name']}` is {ratio:.0%} missing ({col['null']:,} of {col['n']:,}).",
            })

    # High-cardinality categoricals
    for col in profile:
        if col["type"] == "string" and col["non_null"] > 0:
            uniq_ratio = col["unique"] / col["non_null"]
            if uniq_ratio >= 0.95 and col["unique"] >= 20:
                insights.append({
                    "kind": "id-like",
                    "column": col["name"],
                    "text": f"`{col['name']}` looks like an identifier ({col['unique']:,} unique).",
                })
            elif col["unique"] <= 8 and col["non_null"] >= 10:
                top = col.get("top_values", [])
                pretty = ", ".join(f"{t['value']}({t['count']})" for t in top[:3])
                insights.append({
                    "kind": "category",
                    "column": col["name"],
                    "text": f"`{col['name']}` is categorical ({col['unique']} levels): {pretty}.",
                })

    # Numeric outliers / skew
    for col in profile:
        if col["type"] != "number" or not col.get("stats"):
            continue
        st = col["stats"]
        if st["count"] < 5:
            continue
        if st["std"] > 0:
            skew_hint = (st["mean"] - st["median"]) / st["std"]
            if abs(skew_hint) >= 0.5:
                direction = "right-skewed (long tail of large values)" if skew_hint > 0 else "left-skewed (long tail of small values)"
                insights.append({
                    "kind": "skew",
                    "column": col["name"],
                    "text": f"`{col['name']}` is {direction}: mean={st['mean']:.2f}, median={st['median']:.2f}.",
                })
        iqr = st["p75"] - st["p25"]
        if iqr > 0:
            hi = st["p75"] + 1.5 * iqr
            lo = st["p25"] - 1.5 * iqr
            if st["max"] > hi or st["min"] < lo:
                insights.append({
                    "kind": "outliers",
                    "column": col["name"],
                    "text": f"`{col['name']}` has outliers (range {st['min']:.2f} – {st['max']:.2f} vs IQR fences {lo:.2f}/{hi:.2f}).",
                })

    # Time series hint
    date_cols = [c["name"] for c in profile if c["type"] == "date"]
    num_cols = [c["name"] for c in profile if c["type"] == "number"]
    if date_cols and num_cols:
        insights.append({
            "kind": "timeseries",
            "text": f"Detected date column(s) {date_cols} alongside numeric metrics — candidate for trend / time-series chart.",
        })

    return insights


# ─── CLI commands ────────────────────────────────────────────────────────

def cmd_list_files(args: argparse.Namespace) -> int:
    target = Path(args.path).expanduser() if args.path else DATA_DIR
    if not target.exists():
        print(json.dumps({"files": []}))
        return 0
    out = []
    for p in sorted(target.rglob("*")):
        if p.is_file() and p.suffix.lower() in {".csv", ".tsv", ".xlsx", ".xls", ".xlsm"}:
            out.append({
                "path": str(p),
                "name": p.name,
                "size": p.stat().st_size,
                "suffix": p.suffix.lower().lstrip("."),
            })
    print(json.dumps({"files": out}, ensure_ascii=False))
    return 0


def cmd_head(args: argparse.Namespace) -> int:
    header, rows, path = load(args.path, sheet=args.sheet)
    n = min(int(args.n), 20)  # hard ceiling to keep stdout small
    out = {
        "path": str(path),
        "columns": header,
        "rows": rows[:n],
        "total_rows": len(rows),
        "preview_limit": n,
    }
    print(json.dumps(out, ensure_ascii=False))
    return 0


def cmd_profile(args: argparse.Namespace) -> int:
    header, rows, path = load(args.path, sheet=args.sheet)
    cols = profile_columns(header, rows)
    print(json.dumps({"path": str(path), "rows": len(rows), "columns": cols}, ensure_ascii=False))
    return 0


def cmd_describe(args: argparse.Namespace) -> int:
    header, rows, path = load(args.path, sheet=args.sheet)
    cols = profile_columns(header, rows)
    print(f"file:    {path}")
    print(f"shape:   {len(rows):,} rows × {len(header)} columns")
    print()
    for c in cols:
        line = f"  {c['name']:<24} {c['type']:<7} nulls={c['null']:<5} unique={c['unique']}"
        if c["type"] == "number" and c.get("stats", {}).get("count"):
            s = c["stats"]
            line += f"  μ={s['mean']:.2f}  σ={s['std']:.2f}  [{s['min']:.2f} .. {s['max']:.2f}]"
        print(line)
    return 0


def cmd_insights(args: argparse.Namespace) -> int:
    header, rows, path = load(args.path, sheet=args.sheet)
    cols = profile_columns(header, rows)
    ins = derive_insights(header, rows, cols)
    out = {"path": str(path), "rows": len(rows), "insights": ins}
    if args.text:
        print(f"insights for {path}  ({len(rows):,} rows)")
        for i in ins:
            sev = i.get("severity", "")
            tag = f"[{sev}] " if sev else ""
            print(f"  • {tag}{i['text']}")
        return 0
    print(json.dumps(out, ensure_ascii=False))
    return 0


def main(argv: list[str]) -> int:
    parser = argparse.ArgumentParser(description="CSV/Excel profiler + insight extractor.")
    sub = parser.add_subparsers(dest="cmd", required=True)

    p_ls = sub.add_parser("list-files", help="list tabular files under a directory (defaults to module data dir)")
    p_ls.add_argument("--path", help="directory to scan")
    p_ls.set_defaults(fn=cmd_list_files)

    def _add_io(p: argparse.ArgumentParser) -> None:
        p.add_argument("--path", required=True, help="CSV/Excel file (absolute path or relative to module data dir)")
        p.add_argument("--sheet", help="Excel sheet name or 0-based index")

    p_h = sub.add_parser("head", help="first N rows as JSON")
    _add_io(p_h)
    p_h.add_argument("-n", type=int, default=5, help="row cap (hard ceiling 20)")
    p_h.set_defaults(fn=cmd_head)

    p_p = sub.add_parser("profile", help="per-column types + stats as JSON")
    _add_io(p_p)
    p_p.set_defaults(fn=cmd_profile)

    p_d = sub.add_parser("describe", help="human-readable profile")
    _add_io(p_d)
    p_d.set_defaults(fn=cmd_describe)

    p_i = sub.add_parser("insights", help="derived insights (missing, outliers, skew, time-series hints)")
    _add_io(p_i)
    p_i.add_argument("--text", action="store_true", help="emit human-readable bullets instead of JSON")
    p_i.set_defaults(fn=cmd_insights)

    args = parser.parse_args(argv[1:])
    # Normalise sheet: int if numeric.
    if getattr(args, "sheet", None) and args.sheet.isdigit():
        args.sheet = int(args.sheet)
    return args.fn(args)


if __name__ == "__main__":
    raise SystemExit(main(sys.argv))
