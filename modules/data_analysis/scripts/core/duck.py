"""DuckDB convenience layer.

Owns:
  * lazy connect — `pip install duckdb` if missing
  * CSV/XLSX → Parquet ingest (FR-DS-03)
  * register_dataset / register_artifact as queryable views
  * execute_sql with row + column metadata

The PRD specifies DuckDB SQL as the dialect (FR-SQL-04). DuckDB is
PostgreSQL-compatible enough that the same SQL works against a PG-backed
follow-on without modification for the vast majority of queries.
"""

from __future__ import annotations

from pathlib import Path
from typing import Any


def connect():
    try:
        import duckdb  # type: ignore
    except ImportError as exc:
        raise SystemExit(
            "ERROR: duckdb not installed. `pip install duckdb openpyxl` to enable SQL features."
        ) from exc
    return duckdb.connect(database=":memory:")


def connect_warehouse(warehouse_path: Path):
    """Open (or create) the persistent project warehouse DuckDB file."""
    try:
        import duckdb  # type: ignore
    except ImportError as exc:
        raise SystemExit(
            "ERROR: duckdb not installed. `pip install duckdb openpyxl` to enable SQL features."
        ) from exc
    warehouse_path.parent.mkdir(parents=True, exist_ok=True)
    return duckdb.connect(database=str(warehouse_path))


def register_parquet_table(
    warehouse_path: Path, table_name: str, parquet_path: Path
) -> None:
    """Create-or-replace a persistent table in the warehouse from a Parquet file."""
    safe = "".join(ch for ch in table_name if ch.isalnum() or ch == "_") or "t"
    con = connect_warehouse(warehouse_path)
    try:
        p = str(parquet_path).replace("'", "''")
        con.execute(
            f'CREATE OR REPLACE TABLE "{safe}" AS SELECT * FROM read_parquet(\'{p}\')'
        )
    finally:
        con.close()


def drop_warehouse_table(warehouse_path: Path, table_name: str) -> None:
    safe = "".join(ch for ch in table_name if ch.isalnum() or ch == "_")
    if not safe or not warehouse_path.exists():
        return
    con = connect_warehouse(warehouse_path)
    try:
        con.execute(f'DROP TABLE IF EXISTS "{safe}"')
    finally:
        con.close()


def csv_to_parquet(csv_path: Path, parquet_path: Path) -> dict[str, Any]:
    """Use DuckDB to materialise CSV → Parquet. Returns row + col counts."""
    con = connect()
    con.execute(f"CREATE TABLE t AS SELECT * FROM read_csv_auto(?, header=True)", [str(csv_path)])
    con.execute(f"COPY t TO '{parquet_path}' (FORMAT PARQUET)")
    rows = con.execute("SELECT COUNT(*) FROM t").fetchone()[0]
    cols = [r[0] for r in con.execute("DESCRIBE t").fetchall()]
    con.close()
    return {"rows": rows, "columns": cols}


def xlsx_list_sheets(xlsx_path: Path) -> list[str]:
    """Return the ordered sheet-name list of an XLSX/XLSM workbook."""
    try:
        from openpyxl import load_workbook  # type: ignore
    except ImportError as exc:
        raise SystemExit("ERROR: openpyxl not installed (`pip install openpyxl`).") from exc
    wb = load_workbook(filename=str(xlsx_path), read_only=True, data_only=True)
    return list(wb.sheetnames)


def xlsx_to_parquet(xlsx_path: Path, parquet_path: Path, sheet: str | int | None = None) -> dict[str, Any]:
    """Read XLSX via openpyxl into DuckDB, then write Parquet."""
    try:
        from openpyxl import load_workbook  # type: ignore
    except ImportError as exc:
        raise SystemExit("ERROR: openpyxl not installed (`pip install openpyxl`).") from exc

    wb = load_workbook(filename=str(xlsx_path), read_only=True, data_only=True)
    ws = wb.worksheets[sheet] if isinstance(sheet, int) else (wb[sheet] if sheet else wb.active)
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header = [str(c) if c is not None else f"col_{i}" for i, c in enumerate(next(rows_iter))]
    except StopIteration:
        header = []
    data = [tuple(("" if v is None else v) for v in r) for r in rows_iter]

    con = connect()
    # Build a typed temp table via list-of-tuples — DuckDB infers types.
    if not header:
        raise SystemExit("ERROR: empty sheet.")
    col_defs = ", ".join('"' + h + '" VARCHAR' for h in header)
    con.execute(f"CREATE TABLE t ({col_defs})")
    if data:
        placeholders = ",".join(["?"] * len(header))
        con.executemany(f"INSERT INTO t VALUES ({placeholders})", data)
    con.execute(f"COPY t TO '{parquet_path}' (FORMAT PARQUET)")
    rows = con.execute("SELECT COUNT(*) FROM t").fetchone()[0]
    cols = [r[0] for r in con.execute("DESCRIBE t").fetchall()]
    con.close()
    return {"rows": rows, "columns": cols}


def ingest(src: Path, parquet_path: Path, sheet: str | int | None = None) -> dict[str, Any]:
    suffix = src.suffix.lower()
    parquet_path.parent.mkdir(parents=True, exist_ok=True)
    if suffix == ".csv":
        return csv_to_parquet(src, parquet_path)
    if suffix == ".tsv":
        con = connect()
        con.execute("CREATE TABLE t AS SELECT * FROM read_csv_auto(?, header=True, delim='\t')", [str(src)])
        con.execute(f"COPY t TO '{parquet_path}' (FORMAT PARQUET)")
        rows = con.execute("SELECT COUNT(*) FROM t").fetchone()[0]
        cols = [r[0] for r in con.execute("DESCRIBE t").fetchall()]
        con.close()
        return {"rows": rows, "columns": cols}
    if suffix in (".xlsx", ".xlsm", ".xls"):
        return xlsx_to_parquet(src, parquet_path, sheet)
    raise SystemExit(f"ERROR: unsupported source format: {suffix}")


def execute_sql(sql: str, parquet_views: dict[str, Path]) -> dict[str, Any]:
    """Run SQL with each name in `parquet_views` registered as a queryable view.

    Returns: { columns, rows, row_count, dtypes }
    """
    con = connect()
    for name, path in parquet_views.items():
        safe = "".join(ch for ch in name if ch.isalnum() or ch == "_")
        p = str(path).replace("'", "''")
        con.execute(f"CREATE VIEW {safe} AS SELECT * FROM read_parquet('{p}')")
    cur = con.execute(sql)
    columns = [d[0] for d in cur.description] if cur.description else []
    rows = cur.fetchall()
    dtypes = [str(d[1]) for d in (cur.description or [])]
    con.close()
    return {
        "columns": columns,
        "rows": [list(r) for r in rows],
        "row_count": len(rows),
        "dtypes": dtypes,
    }


def profile_parquet(parquet_path: Path, sample_rows: int = 5000) -> dict[str, Any]:
    """Type + null + cardinality profile of a Parquet file (FR-DS-02)."""
    con = connect()
    p = str(parquet_path).replace("'", "''")
    con.execute(f"CREATE VIEW t AS SELECT * FROM read_parquet('{p}')")
    describe = con.execute("DESCRIBE t").fetchall()
    total = con.execute("SELECT COUNT(*) FROM t").fetchone()[0]
    cols: list[dict[str, Any]] = []
    for row in describe:
        name, sql_type = row[0], row[1]
        non_null = con.execute(f'SELECT COUNT("{name}") FROM t').fetchone()[0]
        unique = con.execute(f'SELECT COUNT(DISTINCT "{name}") FROM t').fetchone()[0]
        col: dict[str, Any] = {
            "name": name,
            "type": sql_type,
            "n": total,
            "non_null": int(non_null),
            "null": int(total) - int(non_null),
            "unique": int(unique),
        }
        if sql_type.upper() in ("BIGINT", "INTEGER", "DOUBLE", "FLOAT", "DECIMAL", "HUGEINT", "TINYINT", "SMALLINT") \
                or "DECIMAL" in sql_type.upper() or "DOUBLE" in sql_type.upper():
            stats = con.execute(
                f'SELECT MIN("{name}"), MAX("{name}"), AVG("{name}"), STDDEV_SAMP("{name}"), SUM("{name}") FROM t'
            ).fetchone()
            col["stats"] = {
                "min": _to_py(stats[0]),
                "max": _to_py(stats[1]),
                "mean": _to_py(stats[2]),
                "std": _to_py(stats[3]),
                "sum": _to_py(stats[4]),
            }
        else:
            top = con.execute(
                f'SELECT "{name}", COUNT(*) AS c FROM t WHERE "{name}" IS NOT NULL GROUP BY 1 ORDER BY c DESC LIMIT 5'
            ).fetchall()
            col["top_values"] = [{"value": _to_py(v), "count": int(c)} for v, c in top]
        cols.append(col)
    con.close()
    return {"rows": int(total), "columns": cols}


def _to_py(x: Any) -> Any:
    if x is None:
        return None
    if isinstance(x, (int, float, str, bool)):
        return x
    try:
        return float(x)
    except (TypeError, ValueError):
        return str(x)
