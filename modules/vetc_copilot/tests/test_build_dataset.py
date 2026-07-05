import csv
from pathlib import Path
import openpyxl
from build_dataset import convert


def _wb(tmp: Path) -> str:
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet("Vehicle Dataset")
    ws.append(["vehicle_id", "user_id", "vehicle_type", "inspection_expiry"])
    ws.append(["VEH001", "U001", "Car", "2026-07-20"])
    for name, hdr in [
        ("User Profiles", ["user_id", "user_name"]),
        ("Vehicle Documents", ["document_id", "vehicle_id"]),
        ("Knowledge Dataset", ["knowledge_id", "question"]),
        ("VETC Services", ["service_id", "service_name"]),
        ("Public Evaluation", ["Category", "User ID"]),
    ]:
        s = wb.create_sheet(name)
        s.append(hdr)
        s.append(["X", "Y"])
    p = tmp / "src.xlsm"
    wb.save(p)
    return str(p)


def test_convert_writes_vehicles_csv(tmp_path):
    counts = convert(_wb(tmp_path), str(tmp_path))
    rows = list(csv.DictReader((tmp_path / "data" / "vehicles.csv").open(encoding="utf-8")))
    assert rows[0]["vehicle_id"] == "VEH001"
    assert counts["vehicles"] == 1
