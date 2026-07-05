#!/usr/bin/env python
"""One-time converter: P5 Track-4 .xlsm -> module data/*.csv.

Each source sheet maps to one CSV with a normalized snake_case header. The
module reads only these CSVs at runtime; the workbook stays external.
"""
from __future__ import annotations

import argparse
import csv
from pathlib import Path

# sheet name -> (output csv stem, header first-column marker)
_SHEETS = {
    "User Profiles": "users",
    "Vehicle Dataset": "vehicles",
    "Vehicle Documents": "documents",
    "Knowledge Dataset": "knowledge",
    "VETC Services": "services",
    "Public Evaluation": "eval_scenarios",
}


def _norm(name: str) -> str:
    """Normalize a header cell to snake_case ascii-ish key."""
    return str(name).strip().lower().replace(" ", "_")


def _rows(ws) -> list[list]:
    return [list(r) for r in ws.iter_rows(values_only=True)]


def _dump(ws, path: Path) -> int:
    """Write a sheet's first non-empty header row + data rows to CSV. Returns data row count."""
    rows = [r for r in _rows(ws) if any(c is not None for c in r)]
    if not rows:
        path.write_text("", encoding="utf-8")
        return 0
    header = [_norm(c) for c in rows[0]]
    n = 0
    with path.open("w", newline="", encoding="utf-8") as fh:
        w = csv.writer(fh)
        w.writerow(header)
        for r in rows[1:]:
            w.writerow(["" if c is None else c for c in r])
            n += 1
    return n


def convert(xlsx_path: str, out_dir: str) -> dict[str, int]:
    """Materialize the workbook sheets into ``out_dir/data/*.csv``.

    Args:
        xlsx_path: Path to the P5 ``.xlsm`` workbook.
        out_dir: Module directory; CSVs land in ``out_dir/data``.

    Returns:
        Mapping of output stem to number of data rows written.
    """
    import openpyxl

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    data = Path(out_dir) / "data"
    data.mkdir(parents=True, exist_ok=True)
    counts: dict[str, int] = {}
    for sheet, stem in _SHEETS.items():
        if sheet in wb.sheetnames:
            counts[stem] = _dump(wb[sheet], data / f"{stem}.csv")
    return counts


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(prog="build_dataset", description="P5 xlsm -> data/*.csv")
    parser.add_argument("--xlsx", required=True)
    parser.add_argument("--out", default=str(Path(__file__).resolve().parent.parent))
    args = parser.parse_args(argv)
    print(convert(args.xlsx, args.out))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
