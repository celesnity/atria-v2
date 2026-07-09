#!/usr/bin/env python
"""One-time converter: dataset .xlsx → module file corpus + access data.

Reads the AI Workspace dataset workbook and writes:
  - ``sample_documents/<doc_id>.md`` — front-matter (canonical department_id +
    classification + owner + knowledge_space) followed by the Vietnamese body.
  - ``access/users.csv`` — users with department normalized to department_id.
  - ``access/roles.csv`` and ``access/permissions.csv`` — reference copies.

Department labels differ across sheets (e.g. Documents "HR" vs Users
"Human Resources"); everything is canonicalized to the ``department_id`` from
the Departments sheet so ACL comparisons are exact.
"""
from __future__ import annotations

import argparse
import csv
from pathlib import Path


def _rows(ws) -> list[list]:
    return [list(r) for r in ws.iter_rows(values_only=True)]


def _header_index(rows: list[list], first_col: str) -> int:
    for i, r in enumerate(rows):
        if r and str(r[0]).strip() == first_col:
            return i
    raise ValueError(f"header row starting with {first_col!r} not found")


def _dicts(rows: list[list], header_i: int) -> list[dict]:
    header = [str(c).strip() if c is not None else "" for c in rows[header_i]]
    out: list[dict] = []
    for r in rows[header_i + 1:]:
        if not r or all(c is None for c in r):
            continue
        out.append({header[j]: (r[j] if j < len(r) else None) for j in range(len(header))})
    return out


def build_department_map(dept_rows: list[dict]) -> dict[str, str]:
    """Map every known department label (en/id/vi) to its canonical department_id."""
    dmap: dict[str, str] = {}
    for row in dept_rows:
        dept_id = str(row["department_id"]).strip()
        for key in ("department_id", "department_en", "department_vi"):
            val = row.get(key)
            if val:
                dmap[str(val).strip()] = dept_id
    return dmap


def canonical_department(label: str, dept_map: dict[str, str]) -> str:
    """Return the canonical department_id for a raw label, or the label itself."""
    return dept_map.get(str(label).strip(), str(label).strip())


_KNOWLEDGE_SPACE = {"COMP": "Company Knowledge", "EXEC": "Executive Knowledge"}


def _knowledge_space(dept_id: str) -> str:
    return _KNOWLEDGE_SPACE.get(dept_id, "Department Knowledge")


def _write_document(out_docs: Path, doc: dict, meta: dict, dept_map: dict[str, str]) -> None:
    dept_id = canonical_department(doc["department"], dept_map)
    front = {
        "doc_id": doc["document_id"],
        "title": doc["title"],
        "department": dept_id,
        "classification": doc["classification"],
        "owner": canonical_department(meta.get("owner", doc["department"]), dept_map),
        "knowledge_space": _knowledge_space(dept_id),
        "last_updated": str(meta.get("last_updated", "")),
        "language": str(meta.get("language", "vi")),
    }
    lines = ["---"]
    lines += [f"{k}: {v}" for k, v in front.items()]
    lines += ["---", str(doc["content_vi"] or "").strip(), ""]
    (out_docs / f"{doc['document_id']}.md").write_text("\n".join(lines), encoding="utf-8")


def convert(xlsx_path: str, out_dir: str) -> dict:
    """Materialize the workbook into ``out_dir``. Returns counts."""
    import openpyxl

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    dept_rows = _dicts((r := _rows(wb["Departments"])), _header_index(r, "department_id"))
    dept_map = build_department_map(dept_rows)

    doc_rows = _dicts((r := _rows(wb["Documents"])), _header_index(r, "document_id"))
    meta_rows = _dicts((r := _rows(wb["Document_Metadata"])), _header_index(r, "document_id"))
    meta_by_id = {m["document_id"]: m for m in meta_rows}

    out = Path(out_dir)
    out_docs = out / "sample_documents"
    out_access = out / "access"
    out_docs.mkdir(parents=True, exist_ok=True)
    out_access.mkdir(parents=True, exist_ok=True)

    for doc in doc_rows:
        _write_document(out_docs, doc, meta_by_id.get(doc["document_id"], {}), dept_map)

    user_rows = _dicts((r := _rows(wb["Users"])), _header_index(r, "user_id"))
    with open(out_access / "users.csv", "w", newline="", encoding="utf-8") as fh:
        w = csv.writer(fh)
        w.writerow(["user_id", "full_name", "department", "role", "email", "status"])
        for u in user_rows:
            w.writerow([
                u["user_id"], u.get("full_name", ""),
                canonical_department(u["department"], dept_map),
                u["role"], u.get("email", ""), u.get("status", "Active"),
            ])

    _dump_reference(wb, "Roles", out_access / "roles.csv")
    _dump_reference(wb, "Permissions", out_access / "permissions.csv")

    return {"documents": len(doc_rows), "users": len(user_rows)}


def _dump_reference(wb, sheet: str, path: Path) -> None:
    """Write a sheet's non-empty rows verbatim as a reference CSV."""
    with open(path, "w", newline="", encoding="utf-8") as fh:
        w = csv.writer(fh)
        for r in wb[sheet].iter_rows(values_only=True):
            cells = [c for c in r if c is not None]
            if cells:
                w.writerow(list(r))


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(prog="build_corpus", description="xlsx → module corpus")
    parser.add_argument("--xlsx", required=True, help="Path to the dataset .xlsx.")
    parser.add_argument("--out", default=str(Path(__file__).resolve().parent.parent),
                        help="Module dir to write sample_documents/ and access/ into.")
    args = parser.parse_args(argv)
    counts = convert(args.xlsx, args.out)
    print(counts)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
