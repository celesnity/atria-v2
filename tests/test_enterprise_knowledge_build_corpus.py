# tests/test_enterprise_knowledge_build_corpus.py
"""Converter tests: build a tiny in-memory workbook and assert materialized files."""
from __future__ import annotations

import importlib.util
import sys
from pathlib import Path

_TOOL = (
    Path(__file__).resolve().parent.parent
    / "modules" / "enterprise_knowledge" / "tools" / "build_corpus.py"
)


def _load():
    spec = importlib.util.spec_from_file_location("ek_build_corpus_uut", _TOOL)
    assert spec and spec.loader
    mod = importlib.util.module_from_spec(spec)
    sys.modules["ek_build_corpus_uut"] = mod
    spec.loader.exec_module(mod)
    return mod


def _make_workbook(path):
    import openpyxl
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    dep = wb.create_sheet("Departments")
    dep.append(["Departments"])
    dep.append(["department_id", "department_en", "department_vi", "knowledge_space"])
    dep.append(["HR", "Human Resources", "Nhân sự", "Department Knowledge"])
    dep.append(["COMP", "Company", "Công ty", "Company Knowledge"])
    docs = wb.create_sheet("Documents")
    docs.append(["Enterprise Documents"])
    docs.append([])
    docs.append(["document_id", "title", "department", "classification", "content_vi"])
    docs.append(["DOC007", "Khung lương", "HR", "Confidential", "# Khung lương\nNội dung."])
    meta = wb.create_sheet("Document_Metadata")
    meta.append(["Document Metadata"])
    meta.append([])
    meta.append(["document_id", "title", "department", "classification", "owner",
                 "allowed_access", "last_updated", "tags", "language", "word_count"])
    meta.append(["DOC007", "Khung lương", "HR", "Confidential", "HR",
                 "Own Department", "2025-08-22", "khung, hr", "vi", 100])
    users = wb.create_sheet("Users")
    users.append(["Synthetic Users"])
    users.append([])
    users.append(["user_id", "full_name", "department", "role", "email", "status"])
    users.append(["U001", "Nguyễn Văn An", "Human Resources", "Employee",
                  "u1@synthetic.local", "Active"])
    for name in ("Roles", "Permissions"):
        s = wb.create_sheet(name)
        s.append([name])
        s.append(["col"])
    wb.save(path)


def test_convert_materializes_and_canonicalizes(tmp_path):
    mod = _load()
    xlsx = tmp_path / "ds.xlsx"
    _make_workbook(str(xlsx))
    out = tmp_path / "module"
    counts = mod.convert(str(xlsx), str(out))
    assert counts["documents"] == 1
    # Document front-matter uses canonical department_id HR (already canonical).
    doc_md = (out / "sample_documents" / "DOC007.md").read_text(encoding="utf-8")
    assert "department: HR" in doc_md
    assert "classification: Confidential" in doc_md
    # Users' "Human Resources" is canonicalized to HR.
    users_csv = (out / "access" / "users.csv").read_text(encoding="utf-8")
    assert "U001,Nguyễn Văn An,HR,Employee" in users_csv


def test_canonical_department_maps_hr_alias():
    mod = _load()
    dmap = mod.build_department_map([
        {"department_id": "HR", "department_en": "Human Resources", "department_vi": "Nhân sự"},
    ])
    assert mod.canonical_department("Human Resources", dmap) == "HR"
    assert mod.canonical_department("HR", dmap) == "HR"
